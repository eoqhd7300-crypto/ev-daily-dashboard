"""
build_teardown_data.py
------------------------------------------------------------------
로컬 전용 도구입니다. A2MAC1 xEV Battery Pack teardown 엑셀 원본이 있는 PC에서
직접 실행하세요 (GitHub Actions는 이 폴더에 접근할 수 없습니다).

무엇을 하는가
1) SOURCE_DIR(기본값: D:\\AX data\\AtoMac1 Pack\\xEV Battery Pack) 폴더를 스캔합니다.
2) 파일명 규칙(공백 누락/접두어 중복 등 오탈자 포함)을 정규식 + 보정 로직으로 분석해
   같은 차량의 두 파일 타입을 매칭합니다.
     - Type 2 "... - HV Battery - Data.xlsx"                 -> Cell/Module/Pack 스펙 + 셀 사진
     - Type 1 "... - High Voltage Battery Pack - Data.xlsx"  -> 팩 개요(치수/무게/OEM) + 대표 사진
                                                                  + Navigation 시트 BOM 트리
   (파일명이 "... - Battery Pack - Data.xlsx" 형태인 경우도 Type 1로 처리)
3) 추출한 셀/팩 대표 사진을 assets/teardown/<vehicle-slug>/ 에 저장합니다.
4) 모든 차량 데이터를 teardown_data.json 으로 저장합니다 (대시보드가 fetch로 읽음).

사용법
    python build_teardown_data.py
    python build_teardown_data.py --source "D:\\AX data\\AtoMac1 Pack\\xEV Battery Pack"
    python build_teardown_data.py --only "Tesla Model Y"   # 이름에 포함된 차량만 처리(테스트용)

실행 후 할 일
    git add teardown_data.json assets/teardown
    git commit -m "Update teardown data"
    git push
    (push 하면 deploy_pages 워크플로우가 자동으로 배포합니다)
"""

import argparse
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone

import openpyxl
from PIL import Image as PILImage

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_SOURCE_DIR = r"D:\AX data\AtoMac1 Pack\xEV Battery Pack"
ASSETS_ROOT = os.path.join(BASE_DIR, "assets", "teardown")
OUTPUT_JSON = os.path.join(BASE_DIR, "teardown_data.json")
KST = timezone(timedelta(hours=9))

# Saved images are downscaled + re-compressed to keep the repo lean (see save_image()).
MAX_IMAGE_WIDTH = 480
JPEG_QUALITY = 72

# --- Filename parsing -------------------------------------------------------

TYPE2_RE = re.compile(
    r'^(?P<base>.*?)\s*-\s*(?:xEV Powertrain\s*-\s*)?HV Battery\s*-\s*Data\.xlsx$',
    re.IGNORECASE,
)
TYPE1_RE = re.compile(
    r'^(?P<base>.*?)\s*-\s*(?:High Voltage Battery Pack|Battery Pack)\s*-\s*Data\.xlsx$',
    re.IGNORECASE,
)
YEAR_RE = re.compile(r'^(?P<name>.+?)\s+(?P<year>(?:19|20)\d{2})$')

TOP_VIEW_LABELS = {"Location", "Front", "Back", "Left", "Right", "Top", "Bottom", "Profile"}
# Same set, used when scoping per-part (BOM node) photo extraction under the 'Global' media gallery.
ORIENTATION_LABELS = TOP_VIEW_LABELS
SECTION_HEADER_RE = re.compile(
    r'^(Cell Overview|Cell Characteristics|Module Characteristics.*|Battery Pack Characteristics|'
    r'Energy Densities|Metalwork Properties|System and Function|Media Gallery)$',
    re.IGNORECASE,
)


def normalize_repeated_prefix(base):
    """Fixes filenames where a word-group prefix is accidentally duplicated,
    e.g. 'Land Rover Range Rover Land Rover Range Rover Autobiography...'."""
    words = base.split()
    n = len(words)
    for k in range(1, n // 2 + 1):
        if words[:k] == words[k:2 * k]:
            return ' '.join(words[k:])
    return base


def slugify(text):
    text = text.strip().lower()
    text = re.sub(r'[^a-z0-9]+', '-', text)
    return re.sub(r'-+', '-', text).strip('-')


def clean_value(value):
    """Strips Excel's literal '_x000d_\n' / stray carriage-return artifacts from string cell values."""
    if isinstance(value, str):
        value = value.replace('_x000d_\n', ' / ').replace('_x000d_', ' / ').replace('\r\n', ' / ').replace('\r', ' / ')
        value = re.sub(r'\s*/\s*$', '', value)  # trailing separator left over if the value ended with a break
        return value.strip()
    return value


def split_name_year(base):
    base = re.sub(r'\s+', ' ', base).strip()
    m = YEAR_RE.match(base)
    if m:
        return m.group('name').strip(), int(m.group('year'))
    return base, None


def discover_pairs(source_dir):
    files = [f for f in os.listdir(source_dir) if f.lower().endswith('.xlsx')]
    vehicles = {}  # slug -> {'name','year','type1','type2'}

    for f in files:
        m2 = TYPE2_RE.match(f)
        m1 = None if m2 else TYPE1_RE.match(f)
        if m2:
            base, kind = m2.group('base'), 'type2'
        elif m1:
            base, kind = m1.group('base'), 'type1'
        else:
            print(f'[WARN] Unrecognized filename pattern, skipping: {f}')
            continue

        base = normalize_repeated_prefix(re.sub(r'\s+', ' ', base).strip())
        name, year = split_name_year(base)
        slug = slugify(f'{name}-{year}') if year else slugify(name)

        entry = vehicles.setdefault(slug, {'name': name, 'year': year, 'type1': None, 'type2': None})
        entry[kind] = f

    return vehicles


# --- Generic label/value sheet walker ---------------------------------------

def iter_label_value_rows(ws, max_row=None):
    """Yields (row_index, label, value) for the top-level A/B column pairs,
    skipping pure breadcrumb rows (colA empty, colB contains ' > ')."""
    max_row = max_row or ws.max_row
    for i in range(1, max_row + 1):
        col_a = ws.cell(row=i, column=1).value
        col_b = ws.cell(row=i, column=2).value
        if col_a is None and isinstance(col_b, str) and '>' in col_b:
            continue  # breadcrumb row
        if col_a is None:
            continue
        yield i, col_a, col_b


def find_type1_boundary_row(ws):
    """Finds the row where the top-level pack block ends and the nested
    Sub-Pack/part BOM detail begins (2nd breadcrumb row with >= 2 '>' chars)."""
    for i in range(1, ws.max_row + 1):
        col_b = ws.cell(row=i, column=2).value
        if isinstance(col_b, str) and col_b.count('>') >= 2:
            return i
    return ws.max_row


def enrich_bom_tree_from_overview(ws, bom_tree, asset_dir):
    """Walks the ENTIRE Overview sheet (every nested Sub-Pack/part breadcrumb block, not just
    the top-level pack) and, for each block whose hierarchy path matches a node already present
    in `bom_tree` (i.e. a part actually shown in the BOM table), enriches that node with:
      - Part Code, Width/Height/Depth [mm], Marking (fields the Navigation sheet does not carry)
      - Representative photos: every 'Global' sub-view (Location/Front/Back/Left/Right/Top/
        Bottom/Profile) plus only the FIRST 'Fastener' photo. Measurements/Code/Other views and
        anything else are intentionally skipped to keep the image count bounded.

    Blocks that don't match any bom_tree node (e.g. individual per-instance items like each of
    32 physical cells inside a module, which aren't their own BOM line item) are parsed but
    discarded - no images are saved for them, keeping disk usage scoped to what the UI displays.
    """
    nodes_by_path = {}
    for node in bom_tree:
        path = tuple(lvl for lvl in node['levels'] if lvl)
        nodes_by_path.setdefault(path, node)

    row_to_image = {}
    for img in getattr(ws, '_images', []):
        r = img.anchor._from.row + 1
        c = img.anchor._from.col + 1
        if c == 2:
            row_to_image[r] = img

    current_node = None
    current_fields = {}
    current_fastener_taken = False
    current_view_counts = {}

    def flush():
        if current_node is None:
            return
        current_node['partCode'] = current_fields.get('Part Code')
        current_node['widthMm'] = current_fields.get('Width [mm]')
        current_node['heightMm'] = current_fields.get('Height [mm]')
        current_node['depthMm'] = current_fields.get('Depth [mm]')
        current_node['marking'] = current_fields.get('Marking')

    for i in range(1, ws.max_row + 1):
        col_a = ws.cell(row=i, column=1).value
        col_b = ws.cell(row=i, column=2).value

        if col_a is None and isinstance(col_b, str) and '>' in col_b:
            flush()
            path = tuple(seg.strip() for seg in col_b.split('>'))
            current_node = nodes_by_path.get(path)
            current_fields = {}
            current_fastener_taken = False
            current_view_counts = {}
            continue

        if col_a is None:
            continue
        label = str(col_a).strip()

        if col_b is not None:
            current_fields[label] = clean_value(col_b)
            continue

        # col_b is None here -> either a section header or a photo placeholder row.
        if current_node is None or i not in row_to_image:
            continue

        take_image = False
        if label in ORIENTATION_LABELS:
            take_image = True
        elif label == 'Fastener' and not current_fastener_taken:
            take_image = True
            current_fastener_taken = True

        if not take_image:
            continue

        current_view_counts[label] = current_view_counts.get(label, 0) + 1
        suffix = f'_{current_view_counts[label]}' if current_view_counts[label] > 1 else ''
        file_base = f"part_{slugify(current_node['nodeId'])}_{slugify(label)}{suffix}"
        rel_path = os.path.relpath(save_image(row_to_image[i], asset_dir, file_base), BASE_DIR).replace('\\', '/')
        current_node.setdefault('images', []).append({'view': label, 'path': rel_path})

    flush()


def collect_row_images(ws, rows_of_interest):
    """Maps row index -> saved image bytes/format for images anchored in column B (col idx 1, 0-based)."""
    result = {}
    for img in getattr(ws, '_images', []):
        anchor_row = img.anchor._from.row + 1  # 0-based -> 1-based
        anchor_col = img.anchor._from.col + 1
        if anchor_col == 2 and anchor_row in rows_of_interest:
            result[anchor_row] = img
    return result


def save_image(img, out_dir, base_name):
    """Saves an embedded xlsx image to disk, downscaled to MAX_IMAGE_WIDTH and re-encoded as a
    compressed JPEG to keep the repository lightweight. Falls back to a raw byte dump for the rare
    format Pillow cannot decode."""
    os.makedirs(out_dir, exist_ok=True)
    data = img._data()
    try:
        pil_img = PILImage.open(io.BytesIO(data))
        pil_img.load()
        if pil_img.mode in ('RGBA', 'P', 'LA'):
            pil_img = pil_img.convert('RGB')
        elif pil_img.mode != 'RGB':
            pil_img = pil_img.convert('RGB')
        if pil_img.width > MAX_IMAGE_WIDTH:
            ratio = MAX_IMAGE_WIDTH / pil_img.width
            pil_img = pil_img.resize((MAX_IMAGE_WIDTH, max(1, round(pil_img.height * ratio))), PILImage.LANCZOS)
        out_path = os.path.join(out_dir, f'{base_name}.jpg')
        pil_img.save(out_path, 'JPEG', quality=JPEG_QUALITY, optimize=True)
    except Exception as e:
        print(f'  [WARN] Could not re-compress image ({e}); saving raw bytes instead.')
        ext = (img.format or 'jpeg').lower()
        ext = 'jpg' if ext == 'jpeg' else ext
        out_path = os.path.join(out_dir, f'{base_name}.{ext}')
        with open(out_path, 'wb') as fh:
            fh.write(data)
    return out_path


# --- Type 2 parser (HV Battery - Data.xlsx): Cell / Module(s) / Pack specs ---

def parse_type2(path, slug, asset_dir):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Overview']

    meta = {}
    cell = {}
    modules = []
    pack = {}
    current_module = None
    section = None
    cell_photo_rows = []

    for i, label, value in iter_label_value_rows(ws):
        if SECTION_HEADER_RE.match(str(label).strip()):
            section = label.strip()
            if re.match(r'^Module Characteristics', section, re.IGNORECASE):
                current_module = {}
                modules.append(current_module)
            continue

        if value is None:
            # Photo placeholder (only "Cell" rows appear here in Type 2 files)
            if str(label).strip().lower() == 'cell':
                cell_photo_rows.append(i)
            continue

        value = clean_value(value)
        if section is None:
            meta[label] = value
        elif section.lower() == 'cell characteristics':
            cell[label] = value
        elif section.lower().startswith('module characteristics'):
            if current_module is not None:
                current_module[label] = value
        elif section.lower() == 'battery pack characteristics':
            pack[label] = value
        # 'Energy Densities' section has chart placeholders only -> ignored

    # Extract cell photos
    cell_images = []
    row_to_image = collect_row_images(ws, set(cell_photo_rows))
    for idx, row in enumerate(sorted(row_to_image.keys()), start=1):
        rel_path = os.path.relpath(
            save_image(row_to_image[row], asset_dir, f'cell_{idx}'), BASE_DIR
        ).replace('\\', '/')
        cell_images.append(rel_path)

    return {'meta': meta, 'cell': cell, 'modules': modules, 'pack': pack, 'cellImages': cell_images}


# --- Type 1 parser (High Voltage Battery Pack - Data.xlsx): pack overview + BOM tree ---

def parse_type1(path, slug, asset_dir):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Overview']
    boundary = find_type1_boundary_row(ws)

    pack_overview = {}
    photo_rows = {}  # label -> [row indices]

    for i, label, value in iter_label_value_rows(ws, max_row=boundary - 1):
        label_clean = str(label).strip()
        if SECTION_HEADER_RE.match(label_clean):
            continue
        if value is None:
            if label_clean in TOP_VIEW_LABELS:
                photo_rows.setdefault(label_clean, []).append(i)
            continue
        pack_overview[label_clean] = clean_value(value)

    all_rows_of_interest = {r for rows in photo_rows.values() for r in rows}
    row_to_image = collect_row_images(ws, all_rows_of_interest)

    pack_images = {}
    for label, rows in photo_rows.items():
        key = label.lower()
        saved = []
        for idx, row in enumerate(sorted(rows), start=1):
            if row not in row_to_image:
                continue
            suffix = f'_{idx}' if len(rows) > 1 else ''
            rel_path = os.path.relpath(
                save_image(row_to_image[row], asset_dir, f'pack_{key}{suffix}'), BASE_DIR
            ).replace('\\', '/')
            saved.append(rel_path)
        if not saved:
            continue
        pack_images[key] = saved if len(saved) > 1 else saved[0]

    # BOM tree from the Navigation sheet (flat list; hierarchy is in Part Level 0-5 columns).
    # Column positions are looked up dynamically from the header row rather than hardcoded,
    # since not every vehicle's Navigation sheet uses the exact same column layout.
    bom_tree = []
    if 'Navigation' in wb.sheetnames:
        nav = wb['Navigation']
        header_row = None
        for i in range(1, min(nav.max_row, 15) + 1):
            if nav.cell(row=i, column=1).value == 'Node Id':
                header_row = i
                break
        if header_row:
            col_map = {}
            for c in range(1, nav.max_column + 1):
                header_name = nav.cell(row=header_row, column=c).value
                if header_name:
                    col_map[header_name.strip()] = c

            level_cols = [col_map.get(f'Part Level {n}') for n in range(0, 6)]
            weight_col = col_map.get('Part(s) Weight (kg)')
            material_col = col_map.get('Material Type(s)')
            oem_col = col_map.get('OEM/Tiers')
            partcount_col = col_map.get('Number of Parts')

            for i in range(header_row + 1, nav.max_row + 1):
                node_id = nav.cell(row=i, column=1).value
                if not node_id:
                    continue
                levels = [clean_value(nav.cell(row=i, column=c).value) or '' if c else '' for c in level_cols]
                if not any(levels):
                    continue  # stray footer row (e.g. "Powered by A2MAC1"), not an actual part
                bom_tree.append({
                    'nodeId': node_id,
                    'levels': levels,
                    'weightKg': nav.cell(row=i, column=weight_col).value if weight_col else None,
                    'material': clean_value(nav.cell(row=i, column=material_col).value) if material_col else None,
                    'oemTiers': clean_value(nav.cell(row=i, column=oem_col).value) if oem_col else None,
                    'partCount': nav.cell(row=i, column=partcount_col).value if partcount_col else None,
                })

        # Enrich each BOM node with Part Code + physical size (Width/Height/Depth) + Marking,
        # plus scoped representative photos (all orientation views + first Fastener only),
        # parsed directly from the Overview sheet's per-part breadcrumb blocks. Matched by
        # exact hierarchy path; images are only saved for nodes actually shown in the BOM table.
        enrich_bom_tree_from_overview(ws, bom_tree, asset_dir)

    return {'packOverview': pack_overview, 'packImages': pack_images, 'bomTree': bom_tree}


# --- Main --------------------------------------------------------------------

def build(source_dir, only_filter=None):
    vehicles_map = discover_pairs(source_dir)
    vehicles_out = []

    for slug, info in sorted(vehicles_map.items()):
        if only_filter and only_filter.lower() not in info['name'].lower():
            continue

        print(f"Processing: {info['name']} ({info['year']}) -> {slug}")
        asset_dir = os.path.join(ASSETS_ROOT, slug)

        entry = {
            'id': slug,
            'name': info['name'],
            'year': info['year'],
            'sourceFiles': {'type1': info['type1'], 'type2': info['type2']},
            'meta': {}, 'cell': {}, 'modules': [], 'pack': {}, 'cellImages': [],
            'packOverview': {}, 'packImages': {}, 'bomTree': [],
        }

        if info['type2']:
            try:
                t2 = parse_type2(os.path.join(source_dir, info['type2']), slug, asset_dir)
                entry.update({k: t2[k] for k in ('meta', 'cell', 'modules', 'pack', 'cellImages')})
            except Exception as e:
                print(f'  [ERROR] Failed parsing type2 file: {e}')

        if info['type1']:
            try:
                t1 = parse_type1(os.path.join(source_dir, info['type1']), slug, asset_dir)
                entry.update({k: t1[k] for k in ('packOverview', 'packImages', 'bomTree')})
            except Exception as e:
                print(f'  [ERROR] Failed parsing type1 file: {e}')

        vehicles_out.append(entry)

    return vehicles_out


def main():
    parser = argparse.ArgumentParser(description='Build teardown_data.json from local A2MAC1 xlsx exports.')
    parser.add_argument('--source', default=DEFAULT_SOURCE_DIR, help='Folder containing the xlsx teardown files')
    parser.add_argument('--only', default=None, help='Only process vehicles whose name contains this text (for testing)')
    args = parser.parse_args()

    if not os.path.isdir(args.source):
        print(f'[ERROR] Source folder not found: {args.source}')
        sys.exit(1)

    vehicles = build(args.source, args.only)

    output = {
        'generatedAt': datetime.now(KST).isoformat(),
        'sourceDir': args.source,
        'vehicleCount': len(vehicles),
        'vehicles': vehicles,
    }

    with open(OUTPUT_JSON, 'w', encoding='utf-8') as fh:
        json.dump(output, fh, ensure_ascii=False, indent=2)

    print(f'\nDone. {len(vehicles)} vehicles written to {OUTPUT_JSON}')
    print('Next steps: git add teardown_data.json assets/teardown && git commit && git push')


if __name__ == '__main__':
    main()
